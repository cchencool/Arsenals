import xlrd
import xlwt
import openpyxl


# read xls
def testReadExcel(filename):
    book = xlrd.open_workbook(filename)
    sheet = book.sheet_by_index(0)
    rc = sheet.nrows
    cc = sheet.ncols
    print("work sheet name : %s" % book.sheet_names()[0])
    print("book sheet count : %d" % book.nsheets)
    print("sheet.name : %s, sheet.nrows : %d, sheet.ncols : %d" % (sheet.name, rc, cc))
    for i in range(rc):
        # print("")
        for j in range(cc):
            print("%d:%d-%s" % (i, j, sheet.cell(i, j).value))  # , end='')


# write xls
def testWriteExce(filename):
    book = xlwt.Workbook(filename)
    sheet = book.add_sheet("data2")
    # sheet = book.sheet_index(0)
    sheet.write(0, 0, "test writing")
    row = sheet.row(1)
    row.write(2, 'writing again')

    sheet.flush_row_data()
    book.save(filename)


# read & write xlsx
def testOpenpyxl(filename):
    org_filename = filename

    if not (filename.endswith(".xlsx")):
        filename += ".xlsx"

    wb_in = openpyxl.load_workbook(filename)

    wsNames = wb_in.get_sheet_names()
    # size = len(sheetNames);
    # sheet = wb.get_sheet_by_name(ws[0]);
    ws_in = wb_in[wsNames[0]]
    # ws_in = ws.active

    # value = ws['A1'].value
    # print(value)

    print("Row count : %d " % ws_in.max_row)    # get row count
    print("Column count : %d " % ws_in.max_column) # get column count

    for row in ws_in[1:2]:
        # print(row)
        for cell in row:
            print(cell.value, end='\n')

    # for col in ws_in["A:C"]:
    #     print(col)
    #     for cell in col:
    #         print(cell.value, end="")

    # wb_out = openpyxl.Workbook()
    ws_out = wb_in.create_sheet("new sheet2", 1)

    a = "a"


    ws_out["A1"] = "object"
    ws_out["B1"] = "dir()"

    ws_out["A2"] = "cell"
    ws_out["B2"] = str(dir(ws_out["B2"]));

    print(isinstance(ws_out["A1"], openpyxl.cell.Cell));

    wb_in.save(org_filename + "_New.xlsx")


if __name__ == '__main__':
    filename = "/Users/Chen/Desktop/data"
    # filename2 = "/Users/Chen/Desktop/data.xls"
    # testReadExcel(filename)
    # testWriteExce(filename)
    testOpenpyxl(filename)
